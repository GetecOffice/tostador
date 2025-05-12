from django.shortcuts import render
from .models import *
from django.http import JsonResponse
from django.db.models import Sum
from django.utils import timezone
from datetime import timedelta
import time
# descargar PDF
from datetime import datetime, timedelta
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.templatetags.static import static
from django.utils import timezone
from .models import tblFiltroTostador, tblDatoLineaTostador

from datetime import timedelta, datetime
from django.utils import timezone

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.utils import timezone
from .models import tblFiltroTostador, tblDatoLineaTostador

def archivos(request):
    response = generacion_excel(request)
    time.sleep(1)
    response = generacion_pdf(request)
    return response
def generacion_excel(request):
    # Obtener filtro de la base de datos
    dias = tblFiltroTostador.objects.get(ID=1)
    filtro = dias.Filtro
    ahora = timezone.now()

    # Calcular la fecha de inicio según el filtro
    if filtro == 30:
        fecha_inicio = ahora - timedelta(days=30)
    elif filtro == 7:
        fecha_inicio = ahora - timedelta(days=7)
    elif filtro == 1:
        fecha_inicio = ahora.date()
    else:
        fecha_inicio = None  # Si no es ninguno de los casos anteriores, asignamos None para evitar errores

    # Consultar los datos si la fecha de inicio es válida
    if fecha_inicio:
        datos_tabla = tblDatoLineaTostador.objects.filter(FechaYHora__gte=fecha_inicio).values(
            'ID', 'NoLinea', 'Folio', 'TmpInicial', 'TmpTostado', 'TmpEnfriado',
            'TmpoElevado', 'Operador', 'BatchVerde', 'CantAgua', 'TempCorte', 'PesoTostado', 'Destino', 'FechaYHora'
        )
    else:
        datos_tabla = []  # Si fecha_inicio es None, asignamos una lista vacía
    
    # Crear un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Tostador"
    
    # Establecer encabezados
    headers = ['ID', 'NoLinea', 'Folio', 'TmpInicial', 'TmpTostado', 'TmpEnfriado',
               'TmpoElevado', 'Operador', 'BatchVerde', 'CantAgua', 'TempCorte', 
               'PesoTostado', 'Destino', 'FechaYHora']
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header  # Asignar encabezado en la primera fila
    
    # Llenar los datos en las filas siguientes
    for row_num, row_data in enumerate(datos_tabla, 2):
        ws[f'A{row_num}'] = row_data['ID']
        ws[f'B{row_num}'] = row_data['NoLinea']
        ws[f'C{row_num}'] = row_data['Folio']
        ws[f'D{row_num}'] = row_data['TmpInicial']
        ws[f'E{row_num}'] = row_data['TmpTostado']
        ws[f'F{row_num}'] = row_data['TmpEnfriado']
        ws[f'G{row_num}'] = row_data['TmpoElevado']
        ws[f'H{row_num}'] = row_data['Operador']
        ws[f'I{row_num}'] = row_data['BatchVerde']
        ws[f'J{row_num}'] = row_data['CantAgua']
        ws[f'K{row_num}'] = row_data['TempCorte']
        ws[f'L{row_num}'] = row_data['PesoTostado']
        ws[f'M{row_num}'] = row_data['Destino']
        ws[f'N{row_num}'] = row_data['FechaYHora'].strftime('%Y-%m-%d %H:%M:%S')  # Formatear la fecha

    # Formatear la fecha actual para el nombre del archivo
    fecha_actual = datetime.today()
    formatted_fecha_actual = fecha_actual.strftime("%Y-%m-%d %H-%M-%S")
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Tostador_{formatted_fecha_actual}.xlsx"'

    # Guardar el archivo Excel en la respuesta
    wb.save(response)
    
    return response

def generacion_pdf(request):
    # Obtener filtro de la base de datos
    dias = tblFiltroTostador.objects.get(ID=1)
    filtro = dias.Filtro
    ahora = timezone.now()

    # Calcular la fecha de inicio según el filtro
    if filtro == 30:
        fecha_inicio = ahora - timedelta(days=30)
    elif filtro == 7:
        fecha_inicio = ahora - timedelta(days=7)
    elif filtro == 1:
        fecha_inicio = ahora.date() 
    else:
        fecha_inicio = None  # Si no es ninguno de los casos anteriores, asignamos None para evitar errores

    # Consultar los datos si la fecha de inicio es válida
    if fecha_inicio:
        datos_tabla = tblDatoLineaTostador.objects.filter(FechaYHora__gte=fecha_inicio).values(
            'ID', 'NoLinea', 'Folio', 'TmpInicial', 'TmpTostado', 'TmpEnfriado',
            'TmpoElevado', 'Operador', 'BatchVerde', 'CantAgua', 'TempCorte', 'PesoTostado', 'Destino', 'FechaYHora'
        )
    else:
        datos_tabla = []  # Si fecha_inicio es None, asignamos una lista vacía
        
    # Formateo de la fecha actual para el nombre del archivo
    fecha_actual = datetime.today()
    formatted_fecha_actual = fecha_actual.strftime("%Y-%m-%d %H-%M-%S")

    # Logo y usuario
    logo_url = request.build_absolute_uri(static('assets/img/inicio/valmo.png'))
    user = request.user

    # Renderizar el HTML para el PDF
    html_string = render_to_string('Descargas/PDF/ReporteTostador.html', {
        'logo_url': logo_url, 
        'fecha_actual': fecha_actual, 
        'datos_tabla': datos_tabla,  # Aseguramos que 'datos_tabla' siempre esté definido
        'user': user,
    })

    # Crear un buffer en memoria para el PDF
    pdf_buffer = BytesIO()

    # Generar el PDF usando xhtml2pdf
    pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)

    # Verificar si hubo errores al generar el PDF
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)

    # Obtener el contenido PDF desde el buffer
    pdf_file = pdf_buffer.getvalue()

    # Crear una respuesta HTTP con el archivo PDF adjunto
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Tostador_{formatted_fecha_actual}.pdf"'

    return response

def home(request):
    if request.method == "POST":
        filtro = int(request.POST['filtro'])
        dias = tblFiltroTostador.objects.get(ID=1)
        dias.Filtro = filtro
        dias.save()
    data = tblFiltroTostador.objects.get(ID=1)
    return render(request, 'index.html', {"data":data})
    
def tostador(request):
    if request.method == "POST":
        filtro = int(request.POST['filtro'])
        dias = tblFiltroTostador.objects.get(ID=1)
        dias.Filtro = filtro
        dias.save()
    data = tblFiltroTostador.objects.get(ID=1)
    return render(request, 'Tostador/index.html', {"data":data})

def graficas_y_datos(request):
    dias = tblFiltroTostador.objects.get(ID=1)
    filtro = dias.Filtro

    ahora = timezone.now()

    if filtro == 30:
        fecha_inicio = ahora - timedelta(days=30)
    elif filtro == 7:
        fecha_inicio = ahora - timedelta(days=7)
    elif filtro == 1:
        fecha_inicio = ahora.date() 
    else:
        fecha_inicio = None  

    if fecha_inicio:
        datos_tabla = list(tblDatoLineaTostador.objects.filter(FechaYHora__gte=fecha_inicio).values(
            'ID', 'NoLinea', 'Folio', 'TmpInicial', 'TmpTostado', 'TmpEnfriado',
            'TmpoElevado', 'Operador', 'BatchVerde', 'CantAgua', 'TempCorte', 'PesoTostado', 'Destino', 'FechaYHora'
        ))
        agua = list(tblDatoLineaTostador.objects.filter(FechaYHora__gte=fecha_inicio).values('NoLinea').annotate(total_agua=Sum('CantAgua')))
        peso = list(tblDatoLineaTostador.objects.filter(FechaYHora__gte=fecha_inicio).values('NoLinea').annotate(total_cafe=Sum('PesoTostado')))
        batch = list(tblDatoLineaTostador.objects.filter(FechaYHora__gte=fecha_inicio).values('NoLinea').annotate(total_batch=Sum('BatchVerde')))
        
        graficas_raw = tblDatoLineaTostador.objects.filter(NoLinea__in=['1', '2', '3'], FechaYHora__gte=fecha_inicio).values('TempCorte', 'FechaYHora', 'NoLinea').order_by('-FechaYHora')
    else:
        # Si no se especifica un filtro, puedes traer todos los registros
        datos_tabla = list(tblDatoLineaTostador.objects.values(
            'ID', 'NoLinea', 'Folio', 'TmpInicial', 'TmpTostado', 'TmpEnfriado',
            'TmpoElevado', 'Operador', 'BatchVerde', 'CantAgua', 'TempCorte', 'PesoTostado', 'Destino', 'FechaYHora'
        ))
        agua = list(tblDatoLineaTostador.objects.values('NoLinea').annotate(total_agua=Sum('CantAgua')))
        peso = list(tblDatoLineaTostador.objects.values('NoLinea').annotate(total_cafe=Sum('PesoTostado')))
        batch = list(tblDatoLineaTostador.objects.values('NoLinea').annotate(total_batch=Sum('BatchVerde')))

        graficas_raw = tblDatoLineaTostador.objects.filter(NoLinea__in=['1', '2', '3']).values('TempCorte', 'FechaYHora', 'NoLinea').order_by('-FechaYHora')


    datos_grafica = {'1': [], '2': [], '3': []}
    
    for dato in graficas_raw:
        datos_grafica[dato['NoLinea']].append({
            'hora': dato['FechaYHora'].strftime('%H:%M:%S'),
            'temperatura': dato['TempCorte']
        })
    return JsonResponse({ 'tabla': datos_tabla, 'agua': agua, 'peso': peso, 'batch': batch, 'grafica': datos_grafica}, safe=False)
    
    
    
        
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
